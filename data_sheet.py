#!/usr/bin/python3

import openpyxl as pyxl

main_sheet = "main_sheet.xlsx"
index = "main_hash"
void_client = "XXX"

#Put to deal with pontuatiion
def normalize_the(name):
	aux = name.lower()
	return aux

def create_workbook():
	wb = pyxl.Workbook()
	wb.create_sheet(index)
	ws = wb.get_sheet_by_name(index)
	ws.cell(row=1,column=2,value=0)
	wb.save(main_sheet)


def load_workbook():
	wb = pyxl.load_workbook(main_sheet)
	return wb


#Maybe close, idk
def list_clients():
	cli_list = []
	wb = load_workbook()

	mh = wb.get_sheet_by_name(index)
	nro_clients = int(mh.cell(row=1,column=2).value) + 1

	for i in range(2,nro_clients + 1):
		cli_list.append(mh.cell(row=i,column=2).value)

	return cli_list


#Maybe improve using simmilarity between names
#Problems if two people have the exact same name
#If problems maybe check the limits of rows
#Maybe split in find and add
#The function that calls it should probably save the workbokk(wb)
def find_client(name):
	wb = load_workbook()

	mh = wb.get_sheet_by_name(index)
	cli_identifier = name.lower()
	nro_clients = int(mh.cell(row=1,column=2).value) + 1
	new_sheet = "ws" + str(mh.max_row)

#	print("clientes " + str(nro_clients))

	for i in range(1,nro_clients + 1):
		aux = mh.cell(row=i,column=2).value
#		print("[" + str(i)  + "] " + str(aux))
		if cli_identifier == aux:
			aux = mh.cell(row=i,column=1).value
			return wb.get_sheet_by_name(aux)

	mh.cell(row=1,column=2,value=nro_clients)
	nro_clients += 1
	mh.cell(row=nro_clients,column=1,value=new_sheet)
	mh.cell(row=nro_clients,column=2,value=cli_identifier)
	wb.create_sheet(new_sheet)
	ws = wb.get_sheet_by_name(new_sheet)
	ws.cell(row=1,column=2,value=0)

	return ws
#Maybe referenciation problems because the new sheet is created in this function
def find_client_2(wb,name):

	mh = wb.get_sheet_by_name(index)
	cli_identifier = name.lower()
#	print(cli_identifier)
	nro_clients = int(mh.cell(row=1,column=2).value) + 1
	new_sheet = "ws" + str(mh.max_row)

#	print("clientes " + str(nro_clients))

	for i in range(1,nro_clients + 1):
		aux = mh.cell(row=i,column=2).value
#		print("[" + str(i)  + "] " + str(aux))
		if cli_identifier == aux:
			aux = mh.cell(row=i,column=1).value
			return wb.get_sheet_by_name(aux)

	mh.cell(row=1,column=2,value=nro_clients)
	nro_clients += 1
	mh.cell(row=nro_clients,column=1,value=new_sheet)
	mh.cell(row=nro_clients,column=2,value=cli_identifier)
#	print("WTF!!!")
	wb.create_sheet(new_sheet)
	ws = wb.get_sheet_by_name(new_sheet)
	ws.cell(row=1,column=2,value=0)

	return ws


def list_sessions_by_date(name):
	ws = find_client(name)

	date_list = []
	nro_sessions = int(ws.cell(row=1,column=2).value) + 1
	
	for i in range(2,nro_sessions + 1):
		date_list.append(ws.cell(row=i,column=1).value)

	return date_list


def find_session(name,data):
	ws = find_client(name)
	nro_sessions = int(ws.cell(row=1,column=2).value) + 1

	for i in range(2,nro_sessions + 1):
		if ws.cell(row=i,column=1).value == data:
			return i
	
	return nro_sessions + 1 #Return where to add the new session


def edit_session(nome,line,dt,rt,vp,fq,obs):
	wb = load_workbook()
	ws = find_client_2(wb,nome)

	ws.cell(row=line,column=1,value=dt)
	ws.cell(row=line,column=2,value=rt)
	ws.cell(row=line,column=3,value=vp)
	ws.cell(row=line,column=4,value=fq)
	ws.cell(row=line,column=5,value=obs)

	wb.save(main_sheet)


#Prblems maybe be on the header for session count
def add_session(cliente,dt,rt,vp,fq,obs):
	wb = load_workbook()
	ws = find_client_2(wb,cliente)
	nro_sessions = int(ws.cell(row=1,column=2).value) + 2

	ws.cell(row=nro_sessions,column=1,value=dt)
	ws.cell(row=nro_sessions,column=2,value=rt)
	ws.cell(row=nro_sessions,column=3,value=vp)
	ws.cell(row=nro_sessions,column=4,value=fq)
	ws.cell(row=nro_sessions,column=5,value=obs)

	ws.cell(row=1,column=2,value=nro_sessions)

	wb.save(main_sheet)
#	save_workbook(wb)


#If problems apper maybe use header may fix
def show_all_sessions(wb,cliente):
	ws = find_client(wb,cliente.nome)
	for i in range(2,ws.max_row +1):
		aux = ws.cell(row=i,column=1).value
#		print(aux)

#Need to change the hash value
def edit_cliente(original_name,new_name,new_telf):
	wb = load_workbook()

	if original_name == void_client:
		ws = find_client_2(wb,new_name)		
	else :
		ws = find_client_2(wb,original_name)
	###############################################################
		new_name = normalize_the(new_name)
		original_name = normalize_the(original_name)
		mh = wb.get_sheet_by_name(index)
		nro_clients = int(mh.cell(row=1,column=2).value) + 1
		for i in range(1,nro_clients + 1):
			aux = mh.cell(row=i,column=2).value
			if original_name == aux:
				mh.cell(row=i,column=2,value=new_name)			
	###############################################################

	ws.cell(row=1,column=1,value=new_name)
	ws.cell(row=1,column=3,value=new_telf)

	wb.save(main_sheet)
#def save_workbook(wb):
#	wb.save(main_sheet)
